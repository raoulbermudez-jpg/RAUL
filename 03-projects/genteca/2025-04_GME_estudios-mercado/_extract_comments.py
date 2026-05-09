"""Extrae comentarios abiertos de la encuesta GME, agrupados por aplicación."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from collections import defaultdict

PATH = r'C:\RAUL\03-projects\genteca\2025-04_GME_estudios-mercado\DATA FINAL Pantallas GME.xls.xlsx'
OUT = r'C:\RAUL\03-projects\genteca\2025-04_GME_estudios-mercado\Comentarios_tecnicos_encuesta.md'

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb.active

APP_COL = 6
COMMENT_COL = 19
PREF_COL = 12
STATUS_COL = 5

def short_app(name):
    if not name:
        return 'NA'
    s = str(name).lower()
    if 'motor' in s: return 'Motores'
    if 'bomb' in s: return 'Bombas'
    if 'refrig' in s: return 'Refrigeración'
    return str(name)

by_app = defaultdict(list)
total_complete = 0
total_with_comment = 0

for r in range(2, ws.max_row + 1):
    status = ws.cell(row=r, column=STATUS_COL).value
    if status != 'Participación completa':
        continue
    total_complete += 1
    app = short_app(ws.cell(row=r, column=APP_COL).value)
    pref = ws.cell(row=r, column=PREF_COL).value
    comment = ws.cell(row=r, column=COMMENT_COL).value
    if comment and str(comment).strip():
        total_with_comment += 1
        by_app[app].append({
            'id': ws.cell(row=r, column=1).value,
            'pref': pref,
            'comment': str(comment).strip()
        })

# Build markdown report
lines = []
lines.append('# Comentarios técnicos abiertos — Encuesta GME')
lines.append('')
lines.append('**Fuente:** `DATA FINAL Pantallas GME.xls.xlsx` columna 19 (pregunta 14: "¿Qué otra información considera útil o le gustaría que saliera reflejada en las pantallas de la aplicación móvil?")')
lines.append('')
lines.append(f'**Total respuestas completas:** {total_complete}')
lines.append(f'**Respuestas con comentario abierto:** {total_with_comment}')
lines.append('')
lines.append('Este archivo es input para Vera (feasibility técnica de features pedidas), Vael (mensaje y argumentación de valor) y Producto (decisión de scope para lanzamiento octubre 2026).')
lines.append('')
lines.append('---')
lines.append('')

for app in ['Refrigeración', 'Motores', 'Bombas', 'NA']:
    if app not in by_app:
        continue
    items = by_app[app]
    lines.append(f'## {app} ({len(items)} comentarios)')
    lines.append('')
    for it in items:
        pref = it['pref'] or '(sin preferencia A/B registrada)'
        lines.append(f'- **#{it["id"]}** [pref: {pref}] — {it["comment"]}')
    lines.append('')

lines.append('---')
lines.append('')
lines.append('## Lectura sugerida')
lines.append('')
lines.append('Antes de pasar estos comentarios a Vera/Vael, conviene clasificarlos en 4 categorías temáticas:')
lines.append('')
lines.append('1. **Features adicionales pedidas** (multivoltaje, sensibilidad, picos arranque, historial timestamp, etc.) — input directo Vera para feasibility.')
lines.append('2. **Información que el técnico quiere ver en pantalla** — input directo a equipo de UI/Producto.')
lines.append('3. **Comentarios sobre conectividad / app / experiencia** — input UX/Producto.')
lines.append('4. **Comentarios sobre precio o expectativa de valor** — refuerza/matiza el análisis Van Westendorp.')
lines.append('')
lines.append('Esa clasificación temática NO se hizo aquí — es trabajo posterior con Vera/Vael cuando el equipo lo requiera.')
lines.append('')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Total respuestas completas: {total_complete}')
print(f'Con comentario: {total_with_comment}')
print(f'Por aplicación: {dict((k, len(v)) for k, v in by_app.items())}')
print(f'Archivo: {OUT}')
