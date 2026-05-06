"""Van Westendorp Price Sensitivity Meter — GME survey.
Genera 4 PNG (Total + Motores + Bombas + Refrigeración) + tabla resumen JSON.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
from collections import defaultdict
import json
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np

BASE = Path(r'C:\RAUL\03-projects\genteca\GME Estudios de mercado')
OUT  = BASE / 'van_westendorp'
OUT.mkdir(exist_ok=True)
XLSX = BASE / 'DATA FINAL Pantallas GME.xls.xlsx'

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

def find_col(needle):
    for h in headers:
        if h and needle in h:
            return h
    return None

COL_APP   = find_col('1. Por favor seleccione')
COL_STAT  = find_col('Estado de la participación')
COL_TC    = find_col('TAN BARATO')
COL_C     = find_col('económico y confiable')
COL_E     = find_col('costoso pero aún')
COL_TE    = find_col('excesivamente costoso')

def short_app(name):
    if not name: return 'NA'
    s = name.lower()
    if 'motor' in s: return 'Motores'
    if 'bomb' in s: return 'Bombas'
    if 'refrig' in s: return 'Refrigeración'
    return name

# Cargar respuestas completas
rows = []
for r in range(2, ws.max_row + 1):
    row = {h: ws.cell(row=r, column=c).value for c, h in enumerate(headers, start=1)}
    if row.get(COL_STAT) != 'Participación completa':
        continue
    try:
        tc = float(row[COL_TC]); c = float(row[COL_C])
        e  = float(row[COL_E]);  te = float(row[COL_TE])
    except (TypeError, ValueError):
        continue
    rows.append({'app': short_app(row[COL_APP]),
                 'tc': tc, 'c': c, 'e': e, 'te': te})

print(f'Respuestas válidas para VW: {len(rows)}')

segments = {'Total': rows}
for app in ['Motores', 'Bombas', 'Refrigeración']:
    segments[app] = [r for r in rows if r['app'] == app]
for k, v in segments.items():
    print(f'  {k}: n={len(v)}')

# --- Curvas VW ---
def vw_curves(group, p_max=None, step=0.5):
    """Devuelve (precios, TC, C, E, TE) en porcentajes 0-100."""
    if not group:
        return None
    n = len(group)
    if p_max is None:
        p_max = max(r['te'] for r in group) * 1.15
    P = np.arange(0, p_max + step, step)
    tc = np.array([100.0 * sum(1 for r in group if r['tc'] >= p) / n for p in P])
    c  = np.array([100.0 * sum(1 for r in group if r['c']  >= p) / n for p in P])
    e  = np.array([100.0 * sum(1 for r in group if r['e']  <= p) / n for p in P])
    te = np.array([100.0 * sum(1 for r in group if r['te'] <= p) / n for p in P])
    return P, tc, c, e, te

def first_crossing(P, a, b):
    """Primer precio donde a(P) cruza b(P) (a-b cambia de signo)."""
    diff = a - b
    sign = np.sign(diff)
    # Buscar índice donde el signo cambia (ignorar ceros iniciales)
    for i in range(1, len(P)):
        if sign[i-1] == 0 or sign[i] == 0:
            if sign[i-1] != sign[i]:
                # Interpolación lineal
                if (a[i] - a[i-1]) - (b[i] - b[i-1]) == 0:
                    return P[i]
                t = (b[i-1] - a[i-1]) / ((a[i] - a[i-1]) - (b[i] - b[i-1]))
                return P[i-1] + t * (P[i] - P[i-1])
        elif sign[i-1] != sign[i]:
            t = (b[i-1] - a[i-1]) / ((a[i] - a[i-1]) - (b[i] - b[i-1]))
            return P[i-1] + t * (P[i] - P[i-1])
    return None

results = {}

for seg_name, group in segments.items():
    if not group:
        continue
    curves = vw_curves(group)
    if curves is None:
        continue
    P, tc, c, e, te = curves

    OPP = first_crossing(P, tc, te)   # TC ∩ TE
    IPP = first_crossing(P, c, e)     # C  ∩ E
    PMC = first_crossing(P, tc, e)    # TC ∩ E (lower bound)
    PME = first_crossing(P, c, te)    # C  ∩ TE (upper bound)

    results[seg_name] = {
        'n': len(group),
        'OPP': round(OPP, 2) if OPP else None,
        'IPP': round(IPP, 2) if IPP else None,
        'PMC': round(PMC, 2) if PMC else None,
        'PME': round(PME, 2) if PME else None,
    }
    print(f'\n[{seg_name}] n={len(group)}')
    print(f'  PMC (lower bound):  ${PMC:.2f}' if PMC else '  PMC: n/a')
    print(f'  OPP (optimal):      ${OPP:.2f}' if OPP else '  OPP: n/a')
    print(f'  IPP (indifference): ${IPP:.2f}' if IPP else '  IPP: n/a')
    print(f'  PME (upper bound):  ${PME:.2f}' if PME else '  PME: n/a')
    print(f'  Rango aceptable:    [${PMC:.2f}, ${PME:.2f}]' if PMC and PME else '')

    # --- Plot ---
    # X-axis: recortar a zona relevante (percentil 90 de TE ó max(IPP,PME)*2.5, lo que sea menor)
    te_p90 = float(np.percentile([r['te'] for r in group], 90))
    x_max = min(P.max(), max(te_p90 * 1.05, (PME or 50) * 2.0, 80))

    fig, ax = plt.subplots(figsize=(11, 6.5))
    ax.plot(P, tc, color='#1f77b4', lw=2.2, label='Demasiado barato (TC) ↓')
    ax.plot(P, c,  color='#2ca02c', lw=2.2, label='Económico (C) ↓')
    ax.plot(P, e,  color='#ff7f0e', lw=2.2, label='Costoso (E) ↑')
    ax.plot(P, te, color='#d62728', lw=2.2, label='Excesivo (TE) ↑')

    # Sombrear rango aceptable PRIMERO (debajo de líneas)
    if PMC and PME:
        ax.axvspan(PMC, PME, alpha=0.10, color='#2ca02c', zorder=0)

    # Puntos clave — etiquetas escalonadas para no solaparse
    label_specs = [
        ('PMC', PMC, '#1f77b4', 96, 'Mín. aceptable\n(quality concern)'),
        ('OPP', OPP, '#9467bd', 86, 'Optimal Price\n(min. resistencia)'),
        ('IPP', IPP, '#2ca02c', 76, 'Indifference\n(precio justo)'),
        ('PME', PME, '#d62728', 66, 'Máx. aceptable\n(affordability)'),
    ]
    for label, val, color, ypos, descr in label_specs:
        if val is not None and val <= x_max:
            ax.axvline(val, color=color, ls=':', alpha=0.55, lw=1.3)
            ax.annotate(f'{label}\n${val:.1f}',
                        xy=(val, ypos), xytext=(val, ypos),
                        ha='center', va='center', fontsize=9,
                        color='white', fontweight='bold',
                        bbox=dict(boxstyle='round,pad=0.3',
                                  facecolor=color, edgecolor='none', alpha=0.92))

    # Texto rango aceptable abajo
    if PMC and PME:
        mid = (PMC + PME) / 2
        ax.text(mid, 6, f'Rango aceptable\n[${PMC:.1f} – ${PME:.1f}]',
                ha='center', va='center', fontsize=9,
                color='#1a5e1a', fontweight='bold',
                bbox=dict(boxstyle='round,pad=0.4',
                          facecolor='#d4f0d4', edgecolor='#2ca02c', alpha=0.9))

    ax.set_xlabel('Precio (USD)', fontsize=11, fontweight='bold')
    ax.set_ylabel('% de respondientes', fontsize=11, fontweight='bold')
    ax.set_title(f'Van Westendorp PSM — GME Protector Monofásico — {seg_name} (n={len(group)})',
                 fontsize=12.5, fontweight='bold')
    ax.set_ylim(0, 105)
    ax.set_xlim(0, x_max)
    # Ticks más finos
    if x_max <= 100:
        ax.set_xticks(np.arange(0, x_max + 1, 5))
    else:
        ax.set_xticks(np.arange(0, x_max + 1, 10))
    ax.grid(True, alpha=0.3)
    ax.legend(loc='center right', fontsize=9, framealpha=0.95)
    plt.tight_layout()
    fname = f'vw_{seg_name.lower().replace("ó","o")}.png'
    plt.savefig(OUT / fname, dpi=140, bbox_inches='tight')
    plt.close()
    print(f'  -> {fname}')

# Guardar JSON resumen
with open(OUT / 'vw_summary.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, indent=2, ensure_ascii=False)
print(f'\nResumen JSON: {OUT/"vw_summary.json"}')
