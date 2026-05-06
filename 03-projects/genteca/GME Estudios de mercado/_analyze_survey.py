"""Analiza encuesta DATA FINAL Pantallas GME — segmentación por aplicación."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter, defaultdict
from statistics import mean, median
import openpyxl

PATH = r'C:\RAUL\03-projects\genteca\GME Estudios de mercado\DATA FINAL Pantallas GME.xls.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb.active

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    row = {headers[c-1]: ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)}
    rows.append(row)

# Filtrar completos
def is_complete(r):
    return r.get('Estado de la participación') == 'Participación completa'

complete = [r for r in rows if is_complete(r)]
partial = [r for r in rows if not is_complete(r)]
print(f'Total: {len(rows)} | Completos: {len(complete)} | Parciales: {len(partial)}')
print()

# Mapping app
APP_KEY = '1. Por favor seleccione cual de los siguientes equipos suele instalar con mayor frecuencia'

def short_app(name):
    if not name:
        return 'NA'
    s = name.lower()
    if 'motor' in s: return 'Motores'
    if 'bomb' in s: return 'Bombas'
    if 'refrig' in s: return 'Refrigeración'
    return name

# Segmentos
segs = defaultdict(list)
for r in complete:
    segs[short_app(r[APP_KEY])].append(r)

# Columnas clave
Q_AGRADO_A = '2. B-A -Usando la siguiente escala, por favor indique su nivel de agrado con lo reflejado en la pantalla'
Q_COMPR_A  = '3. B-A- ¿Qué tan comprensible le parece la información que aparece en esta pantalla?'
Q_AGRADO_B = '4. B-B - Usando la siguiente escala, por favor indique su nivel de agrado con lo reflejado en la pantalla'
Q_COMPR_B  = '5. B-B- ¿Qué tan comprensible le parece la información que aparece en esta pantalla?'
Q_UTIL_AB  = '6. B- ¿Cúal de las dos alternativas considera que ofrece información más utíl para su trabajo de instalación y mantenim'
Q_PREF     = '7. B- Si tuviera que seleccionar entre las pantallas A o B, para que sea incluida en la aplicación. ¿Cuál Sería?'
Q_AGRADO_GLOBAL = '8. B- ¿Qué tanto le agrada que la aplicación le ofrezca la información que aparece en las pantallas que acaba de observ'
Q_UTIL_GLOBAL   = '9. B- ¿Qué tan útil le parece la información allí reflejada?'
Q_VW_BARATO     = '10. B- ¿A cuál precio le parece TAN BARATO que no confiaría en su calidad? EXPRESE EL MONTO EN DOLARES'
Q_VW_ECON       = '11. B- ¿A cuál precio le parece económico y confiable?'
Q_VW_COSTOSO    = '12. B- ¿A cuál precio le parece costoso pero aún lo compraría?'
Q_VW_EXCESIVO   = '13. B- ¿A cuál precio le parece excesivamente costoso que No lo compraría ?'
Q_OPEN          = '14. B- ¿Qué otra información considera útil o le gustaría que saliera reflejada en las pantallas de la aplicación móvil'

# Buscar coincidencia parcial (los headers son largos)
def find_col(needle):
    for h in headers:
        if h and needle in h:
            return h
    return None

cols = {
    'agrado_A': find_col('B-A -Usando'),
    'compr_A':  find_col('B-A- ¿Qué tan comprensible'),
    'agrado_B': find_col('B-B - Usando'),
    'compr_B':  find_col('B-B- ¿Qué tan comprensible'),
    'util_ab':  find_col('B- ¿Cúal de las dos'),
    'pref':     find_col('Si tuviera que seleccionar entre las pantallas A o B'),
    'agrado_g': find_col('B- ¿Qué tanto le agrada'),
    'util_g':   find_col('B- ¿Qué tan útil'),
    'vw_barato':   find_col('TAN BARATO'),
    'vw_econ':     find_col('económico y confiable'),
    'vw_costoso':  find_col('costoso pero aún'),
    'vw_excesivo': find_col('excesivamente costoso'),
    'open':        find_col('¿Qué otra información'),
}

print('=== HEADERS ENCONTRADOS ===')
for k, v in cols.items():
    print(f'  {k}: {"OK" if v else "NOT FOUND"}')
print()

def dist(rows, col):
    c = Counter()
    for r in rows:
        v = r.get(col)
        if v is None or v == '':
            continue
        c[str(v).strip()] += 1
    return c

def numstats(rows, col):
    vals = []
    for r in rows:
        v = r.get(col)
        if v is None or v == '':
            continue
        try:
            vals.append(float(v))
        except (TypeError, ValueError):
            pass
    if not vals:
        return None
    vals.sort()
    return {'n': len(vals), 'min': min(vals), 'max': max(vals),
            'mean': round(mean(vals), 2), 'median': median(vals), 'values': vals}

# Imprimir por segmento
def print_seg(name, group):
    print(f'\n========================')
    print(f'SEGMENTO: {name}  (n={len(group)})')
    print(f'========================')
    for label, key in [
        ('Agrado pantalla A',     cols['agrado_A']),
        ('Comprensibilidad A',    cols['compr_A']),
        ('Agrado pantalla B',     cols['agrado_B']),
        ('Comprensibilidad B',    cols['compr_B']),
        ('Más útil (A vs B)',     cols['util_ab']),
        ('Preferencia final',     cols['pref']),
        ('Agrado global app',     cols['agrado_g']),
        ('Utilidad global',       cols['util_g']),
    ]:
        d = dist(group, key)
        print(f'\n  {label}:')
        for k, v in d.most_common():
            print(f'    {k}: {v}')

    print(f'\n  Van Westendorp (USD):')
    for label, key in [
        ('Tan barato (no confío)',  cols['vw_barato']),
        ('Económico y confiable',   cols['vw_econ']),
        ('Costoso pero compraría',  cols['vw_costoso']),
        ('Excesivamente costoso',   cols['vw_excesivo']),
    ]:
        s = numstats(group, key)
        if s:
            print(f'    {label}: n={s["n"]} min={s["min"]} max={s["max"]} '
                  f'media={s["mean"]} mediana={s["median"]}')

    print(f'\n  Comentarios abiertos:')
    for r in group:
        v = r.get(cols['open'])
        if v and str(v).strip().lower() not in ('ninguna', 'ninguno', 'no', 'na', 'n/a', ''):
            print(f'    - "{str(v).strip()}"')

# Global primero
print_seg('TODOS LOS COMPLETOS', complete)
for app in ['Motores', 'Bombas', 'Refrigeración']:
    if app in segs:
        print_seg(app, segs[app])
